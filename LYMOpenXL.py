#-*- coding:utf-8 -*-

import os
from openpyxl import workbook, load_workbook

# 自定义类
class LYMOpenXL:
    def __init__(self, path, read_only=False):
        self.workbook = None
        if os.path.exists(path):
            self.path = path
            self.workbook = load_workbook(self.path,read_only=read_only)
        else:
            print("%s 文件不存在" % path)
            exit(0)

    '''
    获取excel行数
    如果指定表sheet存在，返回行数，否则返回None
    '''
    def get_cell_row(self, sname):
        if self.workbook:
            # 通过工作簿名称获取
            sheet = self.workbook[sname]
            if sheet:
                return sheet.max_row
        return None

    '''
    获取exc列数
    如果指定表sheet存在，返回列数，否则返回None
    '''
    def get_cell_col(self, sname):
        if self.workbook:
            sheet = self.workbook[sname]
            if sheet:
                return sheet.max_column
        return None

    '''
    获取所有工作簿名称列表
    '''
    def get_sheets_name(self):
        if self.workbook:
            return self.workbook.sheetnames
        return None

    '''
    通过索引(第x个表)获取工作簿名称
    索引从0开始
    '''
    def get_sheet_name_by_index(self, index):
        if self.workbook:
            sheets = self.workbook.sheetnames
            wb_len = len(sheets)
            if index>=0 and index<wb_len:
                return sheets[index]
        return None

    '''
    创建工作簿
    '''
    def create_sheet(self, name, index=0):
        res = False
        if self.workbook:
            self.workbook.create_sheet(title=name, index=index)
            res = True
        return res

    '''
    修改sheet名
    '''
    def set_sheet_name(self,sheet,nwe_name):
        res = False
        if self.workbook:
            self.workbook[sheet].title = nwe_name
            res = True

        return res

    '''
    获取单元格值
    '''
    def get_cell_value(self, sheet, row, col):
        value = None
        if self.workbook:
            value = self.workbook[sheet].cell(row=row, column=col).value
        return value

    '''
    修改单元格值
    '''
    def set_cell_vslue(self, sheet, row, col, value):
        res = False
        if self.workbook:
            self.workbook[sheet].cell(row=row, column=col).value = value
            res = True
        return res

    '''
    保存
    '''
    def save(self, path=""):
        if path != "":
            self.path = path
        if self.workbook:
            self.workbook.save(path)

if __name__ == '__main__':
    print("---"*20, end="\n")

    xl = LYMOpenXL("py_excel_emp.xlsx")
    print(">>>获取工作簿列表")
    sheets = xl.get_sheets_name()
    print(sheets)
    print(">>>通过索引获取工作簿名")
    for index in range(0,len(sheets)):
        print(index,": ",xl.get_sheet_name_by_index(index), end='  ')
    print(end="\n")

    print("---" * 20)
    print(">>>从工作簿中获取数据")
    for sheet in sheets:
        nrows = xl.get_cell_row(sheet)
        ncols = xl.get_cell_col(sheet)
        print("---" * 20, end='\n')
        print("工作簿[%s]数据如下：" % sheet)
        for row in range(1,nrows+1):
            for col in range(1, ncols+1):
                value = xl.get_cell_value(sheet,row,col)
                print("[%d, %d]->%s" % (row, col, value), end='\t')

    print("---" * 20)
    print(">>>设置各工作簿中第一行数据为：牛奶")
    for sheet in sheets:
        ncols = xl.get_cell_col(sheet)
        for col in  range(1, ncols+1):
            xl.set_cell_vslue(sheet, row=1, col=col, value="牛奶")

    xl.save("py_excel_emp_new.xlsx")