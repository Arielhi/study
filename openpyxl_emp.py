#-*- coding:utf-8 -*-
from openpyxl import Workbook, load_workbook

if __name__ == '__main__':
    print("python openpyxl实例")

    # 创建excel文档
    wb = Workbook()
    ws = wb.active

    # 给默认的工作簿修改名称
    ws.title = "默认创建的工作簿"
    # 第一行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s1" % col] = "测试数据——水果"
    # 第二行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s2" % col] = "苹果"

    # 创建一个新工作簿
    ws1 = wb.create_sheet("新创建的工作簿1")
    # 第一行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
         ws1["%s1" % col] = "测试数据——蔬菜"
    # 第二行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
         ws1["%s2" % col] = "豆腐"

    # 创建一个新工作簿
    ws1 = wb.create_sheet("新创建的工作簿2")
    # 第一行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
         ws1["%s1" % col] = "测试数据——肉类"
    # 第二行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
         ws1["%s2" % col] = "牛肉"

    # 保存excel文档到硬盘
    wb.save("py_excel_emp.xlsx")

    # 读取excel文件内容
    # 只读模式打开
    read_wb = load_workbook(filename="py_excel_emp.xlsx", read_only=True)
    #获取所有工作簿的名称
    sheets = read_wb.sheetnames

    # 遍历各个工作簿的内容
    # 即第一、二行A-F列的内容
    for sheet in sheets:
        ws = read_wb[sheet]
        print("---"*20)
        print(">>>读取",sheet)
        # 遍历第一、二行A-F列的内容
        for row in (1, 2):
            for col in ("A", "B", "C", "D", "E", "F"):
                print(ws["%s%d" % (col, row)].value, end="   ")
            print(end='\n')

    # 一次性读取多个单元格数据
    print(end="\n\n")
    print(">>>一次性读取多个单元格数据")
    for sheet in sheets:
        ws = read_wb[sheet]
        print("---"*20)
        print(">>>读取", sheet)
        cells_range = ws["A1":"F2"]
        # 遍历已读取的内容
        for cells in cells_range:
            for cell in cells:
                print(cell.value, end=" ")
        print(end='\n')